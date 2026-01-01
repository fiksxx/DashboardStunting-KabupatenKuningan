import streamlit as st
import pandas as pd
import numpy as np
import re
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import geopandas as gpd
import folium
from streamlit_folium import st_folium
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
import os
from PIL import Image
import io
import base64
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches

# Konfigurasi halaman
st.set_page_config(
    page_title="Analisis Data Stunting Kabupaten Kuningan",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# FUNGSI HELPER UNTUK DOWNLOAD GRAFIK
# ============================================================================

def create_download_button_for_chart(fig, filename, title=""):
    """
    Fungsi untuk membuat tombol download grafik Plotly dengan judul
    Menggunakan HTML interaktif (tanpa kaleido) yang kompatibel dengan Streamlit
    
    Parameters:
    - fig: Figure Plotly
    - filename: Nama file output (tanpa ekstensi)
    - title: Judul yang akan ditambahkan di atas grafik
    """
    try:
        import plotly.io as pio
        
        # Buat salinan figure agar tidak mengubah grafik yang ditampilkan
        fig_copy = go.Figure(fig)
        
        # Tambahkan judul jika ada
        if title:
            fig_copy.update_layout(
                title=dict(
                    text=title,
                    font=dict(size=24, family='Poppins', color='#667eea'),
                    x=0.5,
                    xanchor='center'
                )
            )
        
        # Tambahkan background putih untuk tampilan yang bersih
        fig_copy.update_layout(
            paper_bgcolor='white',
            plot_bgcolor='white',
            font=dict(color='#1a1a1a'),
            width=1600,
            height=1000
        )
        
        # Konversi ke HTML interaktif yang bisa dibuka di browser
        html_string = pio.to_html(
            fig_copy, 
            include_plotlyjs='cdn',
            config={
                'toImageButtonOptions': {
                    'format': 'png',
                    'filename': filename,
                    'height': 1000,
                    'width': 1600,
                    'scale': 2
                },
                'displayModeBar': True,
                'displaylogo': False,
                'modeBarButtonsToAdd': ['downloadImage']
            }
        )
        
        # Encode ke base64
        html_bytes = html_string.encode()
        b64 = base64.b64encode(html_bytes).decode()
        
        # Buat tombol download HTML
        href = f'<a href="data:text/html;base64,{b64}" download="{filename}.html" style="text-decoration: none;">' \
               f'<button style="' \
               f'background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);' \
               f'color: white;' \
               f'border: none;' \
               f'border-radius: 12px;' \
               f'padding: 0.8rem 1.5rem;' \
               f'font-weight: 600;' \
               f'font-size: 1rem;' \
               f'font-family: Poppins, sans-serif;' \
               f'cursor: pointer;' \
               f'box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);' \
               f'transition: all 0.3s ease;' \
               f'" ' \
               f'onmouseover="this.style.transform=\'translateY(-2px)\'; this.style.boxShadow=\'0 6px 20px rgba(102, 126, 234, 0.4)\';" ' \
               f'onmouseout="this.style.transform=\'translateY(0)\'; this.style.boxShadow=\'0 4px 12px rgba(102, 126, 234, 0.3)\';">' \
               f'📥 Download Grafik (HTML)</button></a>'
        
        st.markdown(href, unsafe_allow_html=True)
        st.caption("💡 Buka file HTML di browser, lalu gunakan tombol 📷 (camera) untuk download PNG")
        
    except Exception as e:
        st.error(f"⚠️ Tidak dapat membuat tombol download: {str(e)}")
        st.info("💡 Tip: Gunakan tombol kamera 📷 di pojok kanan atas grafik untuk screenshot manual")

def create_static_map_image(data_gdf_merged, title="Peta Sebaran Stunting Per Desa"):
    """
    Fungsi untuk membuat peta statis menggunakan matplotlib yang bisa didownload
    
    Parameters:
    - data_gdf_merged: GeoDataFrame yang sudah di-merge dengan data stunting
    - title: Judul peta
    
    Returns:
    - img_bytes: Image dalam format bytes
    """
    try:
        # Buat figure dengan size besar
        fig, ax = plt.subplots(1, 1, figsize=(20, 16))
        
        # Fungsi warna
        def get_color(persen_stunting):
            if persen_stunting == 0:
                return '#e0e0e0'
            elif persen_stunting < 5:
                return '#fff3cd'
            elif persen_stunting < 10:
                return '#ffcc80'
            elif persen_stunting < 15:
                return '#ff8c42'
            elif persen_stunting < 20:
                return '#ff6b6b'
            else:
                return '#d9534f'
        
        # Plot peta
        data_gdf_merged['color'] = data_gdf_merged['persen_stunting'].apply(get_color)
        data_gdf_merged.plot(
            ax=ax,
            color=data_gdf_merged['color'],
            edgecolor='#34495e',
            linewidth=0.5
        )
        
        # Tambahkan label untuk setiap desa
        for idx, row in data_gdf_merged.iterrows():
            if row['persen_stunting'] > 0:
                centroid = row['geometry'].centroid
                ax.annotate(
                    text=f"{row['NAMOBJ']}\n{row['persen_stunting']:.1f}%",
                    xy=(centroid.x, centroid.y),
                    fontsize=6,
                    ha='center',
                    va='center',
                    bbox=dict(boxstyle='round,pad=0.3', facecolor='white', alpha=0.7, edgecolor='none')
                )
        
        # Styling
        ax.set_title(title, fontsize=28, fontweight='bold', color='#667eea', pad=20, fontfamily='sans-serif')
        ax.axis('off')
        
        # Legend
        legend_elements = [
            mpatches.Patch(facecolor='#e0e0e0', edgecolor='#ccc', label='Tidak ada data'),
            mpatches.Patch(facecolor='#fff3cd', edgecolor='#ffeeba', label='< 5% (Sangat Rendah)'),
            mpatches.Patch(facecolor='#ffcc80', edgecolor='#ffb84d', label='5-10% (Rendah)'),
            mpatches.Patch(facecolor='#ff8c42', edgecolor='#ff7700', label='10-15% (Sedang)'),
            mpatches.Patch(facecolor='#ff6b6b', edgecolor='#ff5555', label='15-20% (Tinggi)'),
            mpatches.Patch(facecolor='#d9534f', edgecolor='#c9302c', label='> 20% (Sangat Tinggi)')
        ]
        
        ax.legend(
            handles=legend_elements,
            loc='lower left',
            fontsize=12,
            title='Prevalensi Stunting',
            title_fontsize=14,
            frameon=True,
            fancybox=True,
            shadow=True
        )
        
        # Tambahkan watermark
        fig.text(0.99, 0.01, 'Dinas Kesehatan Kabupaten Kuningan', 
                ha='right', va='bottom', fontsize=10, color='gray', alpha=0.7)
        
        plt.tight_layout()
        
        # Konversi ke bytes
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=300, bbox_inches='tight', facecolor='white')
        buf.seek(0)
        img_bytes = buf.read()
        plt.close(fig)
        
        return img_bytes
        
    except Exception as e:
        st.error(f"⚠️ Error membuat peta statis: {str(e)}")
        return None

def create_download_button_for_map(img_bytes, filename):
    """
    Fungsi untuk membuat tombol download peta
    
    Parameters:
    - img_bytes: Image dalam format bytes
    - filename: Nama file output (tanpa ekstensi)
    """
    if img_bytes:
        b64 = base64.b64encode(img_bytes).decode()
        
        href = f'<a href="data:image/png;base64,{b64}" download="{filename}.png" style="text-decoration: none;">' \
               f'<button style="' \
               f'background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);' \
               f'color: white;' \
               f'border: none;' \
               f'border-radius: 12px;' \
               f'padding: 0.8rem 1.5rem;' \
               f'font-weight: 600;' \
               f'font-size: 1rem;' \
               f'font-family: Poppins, sans-serif;' \
               f'cursor: pointer;' \
               f'box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);' \
               f'transition: all 0.3s ease;' \
               f'" ' \
               f'onmouseover="this.style.transform=\'translateY(-2px)\'; this.style.boxShadow=\'0 6px 20px rgba(102, 126, 234, 0.4)\';" ' \
               f'onmouseout="this.style.transform=\'translateY(0)\'; this.style.boxShadow=\'0 4px 12px rgba(102, 126, 234, 0.3)\';">' \
               f'📥 Download Peta (PNG Resolusi Tinggi)</button></a>'
        
        st.markdown(href, unsafe_allow_html=True)

# Custom CSS - DIPERCANTIK
st.markdown("""
<style>
    /* Import Google Fonts */
    @import url('https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;600;700&display=swap');
    
    /* Global Styling */
    * {
        font-family: 'Poppins', sans-serif;
    }
    
    /* Main Header dengan Gradient */
    .main-header {
        font-size: 3.5rem;
        font-weight: 700;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        text-align: center;
        margin-bottom: 0.5rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
        animation: fadeInDown 1s ease-in-out;
    }
    
    .sub-header {
        font-size: 1.6rem;
        color: #5f6368;
        text-align: center;
        margin-bottom: 2rem;
        font-weight: 400;
        animation: fadeInUp 1s ease-in-out;
    }
    
    /* Animasi */
    @keyframes fadeInDown {
        from {
            opacity: 0;
            transform: translateY(-20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    /* Metric Cards */
    div[data-testid="stMetricValue"] {
        font-size: 2.2rem;
        font-weight: 700;
        color: #1a1a1a;
    }
    
    div[data-testid="stMetricLabel"] {
        font-size: 1rem;
        font-weight: 500;
        color: #5f6368;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }
    
    div[data-testid="stMetricDelta"] {
        font-size: 1rem;
        font-weight: 600;
    }
    
    /* Styling untuk metric container */
    [data-testid="stMetric"] {
        background: linear-gradient(145deg, #ffffff 0%, #f8f9fa 100%);
        padding: 1.8rem 1.5rem;
        border-radius: 16px;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.08);
        border: 1px solid #e8eaed;
        transition: all 0.3s ease;
    }
    
    [data-testid="stMetric"]:hover {
        transform: translateY(-5px);
        box-shadow: 0 8px 25px rgba(102, 126, 234, 0.15);
        border-color: #667eea;
    }
    
    /* Tabs Styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 1rem;
        background: linear-gradient(145deg, #f8f9fa 0%, #ffffff 100%);
        padding: 0.8rem 1rem;
        border-radius: 16px;
        box-shadow: 0 2px 10px rgba(0, 0, 0, 0.05);
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 3.5rem;
        padding: 0 2rem;
        font-size: 1.05rem;
        font-weight: 600;
        background-color: transparent;
        border-radius: 12px;
        color: #5f6368;
        transition: all 0.3s ease;
        border: 2px solid transparent;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background: linear-gradient(135deg, #667eea15 0%, #764ba215 100%);
        color: #667eea;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
        color: white !important;
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
    }
    
    /* Info Box dengan Border Gradient */
    .info-box {
        background: linear-gradient(145deg, #ffffff 0%, #f8f9fa 100%);
        padding: 1.5rem;
        border-radius: 16px;
        border-left: 5px solid #667eea;
        margin: 1.5rem 0;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.08);
        transition: all 0.3s ease;
    }
    
    .info-box:hover {
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.15);
        transform: translateX(5px);
    }
    
    .info-box b {
        color: #667eea;
        font-weight: 600;
    }
    
    /* Feature Cards */
    .feature-card {
        background: linear-gradient(145deg, #ffffff 0%, #f8f9fa 100%);
        padding: 2rem;
        border-radius: 16px;
        box-shadow: 0 4px 20px rgba(0, 0, 0, 0.08);
        border: 1px solid #e8eaed;
        transition: all 0.3s ease;
        height: 100%;
    }
    
    .feature-card:hover {
        transform: translateY(-8px);
        box-shadow: 0 8px 30px rgba(102, 126, 234, 0.2);
        border-color: #667eea;
    }
    
    .feature-card h4 {
        color: #667eea;
        font-weight: 700;
        font-size: 1.3rem;
        margin-bottom: 1rem;
    }
    
    /* Sidebar Styling */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
        padding: 2rem 1rem;
    }
    
    [data-testid="stSidebar"] .stMarkdown {
        color: white;
    }
    
    [data-testid="stSidebar"] h3 {
        color: white !important;
        font-weight: 700;
        text-align: center;
    }
    
    [data-testid="stSidebar"] .stButton button {
        background: white;
        color: #667eea;
        border: none;
        border-radius: 10px;
        font-weight: 600;
        padding: 0.5rem 1rem;
        transition: all 0.3s ease;
    }
    
    [data-testid="stSidebar"] .stButton button:hover {
        background: #f0f0f0;
        box-shadow: 0 4px 12px rgba(0, 0, 0, 0.15);
    }
    
    /* File Uploader */
    [data-testid="stFileUploader"] {
        background: rgba(255, 255, 255, 0.1);
        border: 2px dashed rgba(255, 255, 255, 0.3);
        border-radius: 12px;
        padding: 1rem;
        transition: all 0.3s ease;
    }
    
    [data-testid="stFileUploader"]:hover {
        border-color: rgba(255, 255, 255, 0.6);
        background: rgba(255, 255, 255, 0.15);
    }
    
    /* Download Buttons */
    .stDownloadButton button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.8rem 1.5rem;
        font-weight: 600;
        font-size: 1rem;
        transition: all 0.3s ease;
        box-shadow: 0 4px 12px rgba(102, 126, 234, 0.3);
    }
    
    .stDownloadButton button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.4);
    }
    
    /* Data Frame Styling */
    [data-testid="stDataFrame"] {
        border-radius: 12px;
        overflow: hidden;
        box-shadow: 0 4px 15px rgba(0, 0, 0, 0.08);
    }
    
    /* Expander */
    .streamlit-expanderHeader {
        background: linear-gradient(145deg, #f8f9fa 0%, #ffffff 100%);
        border-radius: 12px;
        font-weight: 600;
        color: #667eea;
        border: 1px solid #e8eaed;
    }
    
    .streamlit-expanderHeader:hover {
        background: linear-gradient(135deg, #667eea15 0%, #764ba215 100%);
    }
    
    /* Alert Boxes */
    .stAlert {
        border-radius: 12px;
        border-left: 5px solid;
        padding: 1rem 1.5rem;
        box-shadow: 0 2px 10px rgba(0, 0, 0, 0.05);
    }
    
    /* Radio Buttons */
    .stRadio > label {
        font-weight: 600;
        color: #1a1a1a;
        font-size: 1rem;
    }
    
    /* Slider */
    .stSlider > label {
        font-weight: 600;
        color: #1a1a1a;
        font-size: 1rem;
    }
    
    /* Select Box */
    .stSelectbox > label {
        font-weight: 600;
        color: #1a1a1a;
        font-size: 1rem;
    }
    
    /* Text Input */
    .stTextInput > label {
        font-weight: 600;
        color: #1a1a1a;
        font-size: 1rem;
    }
    
    /* Divider */
    hr {
        margin: 2rem 0;
        border: none;
        height: 2px;
        background: linear-gradient(90deg, transparent, #667eea, transparent);
    }
    
    /* Footer Card */
    .footer-card {
        text-align: center;
        padding: 2.5rem;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 20px;
        color: white;
        margin-top: 3rem;
        box-shadow: 0 8px 30px rgba(102, 126, 234, 0.3);
        animation: fadeIn 1s ease-in-out;
    }
    
    .footer-card h3 {
        margin: 0;
        font-size: 1.8rem;
        font-weight: 700;
    }
    
    .footer-card p {
        font-size: 1.1rem;
        margin: 0.5rem 0;
        font-weight: 400;
    }
    
    @keyframes fadeIn {
        from { opacity: 0; }
        to { opacity: 1; }
    }
    
    /* Scrollbar */
    ::-webkit-scrollbar {
        width: 10px;
        height: 10px;
    }
    
    ::-webkit-scrollbar-track {
        background: #f1f1f1;
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(135deg, #5568d3 0%, #6a3f8f 100%);
    }
    
    /* Logo Header Container */
    .logo-header-container {
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 2rem;
        margin-bottom: 2rem;
    }
    
    .logo-img {
        width: 100px;
        height: 100px;
        object-fit: contain;
        animation: fadeIn 1s ease-in-out;
    }
</style>
""", unsafe_allow_html=True)

# ============================================================================
# FUNGSI ETL BARU (MENGGANTIKAN YANG LAMA)
# ============================================================================

def convert_html_xls_to_xlsx(input_path, output_path):
    """Konversi file XLS HTML ke XLSX"""
    try:
        with open(input_path, "r", encoding="utf-8", errors="ignore") as f:
            html = f.read()
        soup = BeautifulSoup(html, "html.parser")
        
        raw_text = soup.get_text("\n", strip=True)
        match = re.search(r'Data Tanggal\s*:\s*([0-9:\-\s]+)', raw_text)
        tanggal_info = match.group(0) if match else ""

        table = soup.find("table")
        if not table:
            return input_path

        wb = Workbook()
        ws = wb.active
        
        row_idx = 1
        if tanggal_info:
            ws.cell(row=row_idx, column=1).value = tanggal_info
            row_idx += 2 

        for row in table.find_all("tr"):
            col_idx = 1
            for cell in row.find_all(["td", "th"]):
                txt = cell.get_text(strip=True)
                colspan = int(cell.get("colspan", 1))
                rowspan = int(cell.get("rowspan", 1))
                
                while isinstance(ws.cell(row=row_idx, column=col_idx), MergedCell):
                    col_idx += 1
                
                ws.cell(row=row_idx, column=col_idx).value = txt
                if colspan > 1 or rowspan > 1:
                    ws.merge_cells(start_row=row_idx, start_column=col_idx, 
                                   end_row=row_idx+rowspan-1, end_column=col_idx+colspan-1)
                col_idx += colspan
            row_idx += 1
            
        wb.save(output_path)
        return output_path
    except Exception as e:
        return input_path

def ensure_xlsx(file_path):
    """Pastikan file dalam format XLSX"""
    root, ext = os.path.splitext(file_path)
    if ext.lower() == ".xls":
        output_path = root + ".xlsx"
        if os.path.exists(output_path):
            return output_path
        return convert_html_xls_to_xlsx(file_path, output_path)
    return file_path

def clean_dataframe(df, col_name_check):
    """Membersihkan baris kosong, baris 'Jumlah', dan baris sampah"""
    df = df.dropna(subset=[col_name_check])
    df = df[~df[col_name_check].astype(str).str.contains("Jumlah|Total|Puskesmas|No", case=False, na=False)]
    return df

def clean_name(text):
    """Membersihkan nama wilayah"""
    return str(text).strip().upper() if pd.notnull(text) else ""

def safe_to_numeric(series):
    """Konversi ke numeric dengan aman"""
    return pd.to_numeric(series, errors='coerce').fillna(0)

def proses_etl(file_gizi, file_sasaran):
    """
    Proses ETL dengan kode baru yang menggunakan 2 file input:
    - file_gizi: File status gizi
    - file_sasaran: File sasaran balita
    """
    try:
        # Konversi file jika diperlukan
        real_file_gizi = ensure_xlsx(file_gizi)
        real_file_sasaran = ensure_xlsx(file_sasaran)
        
        # 1. DIMENSI WAKTU
        df_time = pd.read_excel(real_file_gizi, nrows=1, header=None)
        time_str = str(df_time.iloc[0, 0])
        match = re.search(r'(\d{4})-(\d{2})-(\d{2})\s+(\d{2}):(\d{2}):(\d{2})', time_str)
        
        if match:
            tahun, bulan_num, tanggal, jam, menit, _ = map(int, match.groups())
            bulan_map = {
                1:'JANUARI', 2:'FEBRUARI', 3:'MARET', 4:'APRIL', 5:'MEI', 6:'JUNI',
                7:'JULI', 8:'AGUSTUS', 9:'SEPTEMBER', 10:'OKTOBER', 11:'NOVEMBER', 12:'DESEMBER'
            }
            bulan_str = bulan_map.get(bulan_num, 'UNKNOWN')
        else:
            tahun, bulan_str, tanggal, jam, menit = 2025, 'UNKNOWN', 1, 0, 0
        
        df_waktu = pd.DataFrame([{
            'id_waktu': 1, 'tahun': tahun, 'bulan': bulan_str, 
            'tanggal': tanggal, 'jam': jam, 'menit': menit
        }])
        
        # 2. PROSES STATUS GIZI
        df_gizi = pd.read_excel(real_file_gizi, skiprows=3, header=None)
        
        cols_gizi = [
            'no', 'puskesmas', 'desa',
            'bbu_sangat_kurang', 'bbu_kurang', 'bbu_normal', 'bbu_risiko_lebih', 'bbu_outlier',
            'tbu_sangat_pendek', 'tbu_pendek', 'tbu_normal', 'tbu_tinggi', 'tbu_outlier',
            'bbtb_gizi_buruk', 'bbtb_gizi_kurang', 'bbtb_normal', 'bbtb_risiko_gizi_lebih', 
            'bbtb_gizi_lebih', 'bbtb_obesitas'
        ]
        df_gizi = df_gizi.iloc[:, :len(cols_gizi)]
        df_gizi.columns = cols_gizi
        
        # Bersihkan data
        df_gizi = clean_dataframe(df_gizi, 'puskesmas')
        
        df_gizi['puskesmas_clean'] = df_gizi['puskesmas'].apply(clean_name)
        df_gizi['desa_clean'] = df_gizi['desa'].apply(clean_name)
        df_gizi['join_key'] = df_gizi['puskesmas_clean'] + "_" + df_gizi['desa_clean']
        
        # Konversi angka & hitung
        for col in cols_gizi[3:]:
            df_gizi[col] = safe_to_numeric(df_gizi[col])
        
        df_gizi['jumlah_ditimbang_d'] = df_gizi[['bbu_sangat_kurang', 'bbu_kurang', 'bbu_normal', 'bbu_risiko_lebih', 'bbu_outlier']].sum(axis=1)
        df_gizi['jumlah_kurang_gizi'] = df_gizi['bbu_sangat_kurang'] + df_gizi['bbu_kurang']
        df_gizi['jumlah_stunting'] = df_gizi['tbu_sangat_pendek'] + df_gizi['tbu_pendek']
        df_gizi['jumlah_wasting'] = df_gizi['bbtb_gizi_buruk'] + df_gizi['bbtb_gizi_kurang']
        
        # 3. PROSES SASARAN BALITA
        df_sasaran = pd.read_excel(real_file_sasaran, skiprows=3, header=None)
        df_sasaran = df_sasaran.iloc[:, :6]
        df_sasaran.columns = ['no', 'puskesmas', 'desa', 'sasaran_laki', 'sasaran_perempuan', 'sasaran_total']
        
        df_sasaran = clean_dataframe(df_sasaran, 'puskesmas')
        
        df_sasaran['puskesmas_clean'] = df_sasaran['puskesmas'].apply(clean_name)
        df_sasaran['desa_clean'] = df_sasaran['desa'].apply(clean_name)
        df_sasaran['join_key'] = df_sasaran['puskesmas_clean'] + "_" + df_sasaran['desa_clean']
        
        for col in ['sasaran_laki', 'sasaran_perempuan', 'sasaran_total']:
            df_sasaran[col] = safe_to_numeric(df_sasaran[col])
        
        df_sasaran_join = df_sasaran[['join_key', 'sasaran_laki', 'sasaran_perempuan', 'sasaran_total']]
        
        # 4. GABUNG DATA
        df_gabung = pd.merge(df_gizi, df_sasaran_join, on='join_key', how='left')
        df_gabung['sasaran_total'] = df_gabung['sasaran_total'].fillna(0)
        
        # Hitung persentase
        def calc_percent(num, denom):
            return (num / denom.replace(0, 1)) * 100
        
        df_gabung['persentase_ds'] = calc_percent(df_gabung['jumlah_ditimbang_d'], df_gabung['sasaran_total'])
        df_gabung['persen_kurang_gizi'] = calc_percent(df_gabung['jumlah_kurang_gizi'], df_gabung['jumlah_ditimbang_d'])
        df_gabung['persen_stunting'] = calc_percent(df_gabung['jumlah_stunting'], df_gabung['jumlah_ditimbang_d'])
        df_gabung['persen_wasting'] = calc_percent(df_gabung['jumlah_wasting'], df_gabung['jumlah_ditimbang_d'])
        
        df_gabung = df_gabung[df_gabung['join_key'] != "_"]
        
        # 5. DIMENSI WILAYAH
        df_wilayah = df_gabung[['puskesmas', 'desa']].drop_duplicates().reset_index(drop=True)
        df_wilayah.insert(0, 'id_wilayah', range(1, 1 + len(df_wilayah)))
        
        # 6. FACT TABLE
        df_fact = pd.merge(df_gabung, df_wilayah, on=['puskesmas', 'desa'], how='left')
        df_fact['id_waktu'] = 1
        
        cols_final = [
            'id_wilayah', 'id_waktu', 'puskesmas', 'desa',
            'sasaran_total', 'sasaran_laki', 'sasaran_perempuan',
            'jumlah_ditimbang_d', 'persentase_ds',
            'jumlah_kurang_gizi', 'persen_kurang_gizi',
            'jumlah_stunting', 'persen_stunting',
            'jumlah_wasting', 'persen_wasting',
            'bbu_sangat_kurang', 'bbu_kurang',
            'tbu_sangat_pendek', 'tbu_pendek',
            'bbtb_gizi_buruk', 'bbtb_gizi_kurang', 'bbtb_obesitas'
        ]
        
        df_fact_final = df_fact[cols_final]
        
        return df_fact_final, df_wilayah, df_waktu, True, "Proses ETL berhasil!"
    
    except Exception as e:
        return None, None, None, False, f"Error: {str(e)}"

# ============================================================================
# FUNGSI LOAD SHAPEFILE
# ============================================================================

@st.cache_data
def load_shapefile(shp_path):
    """Load shapefile untuk peta"""
    try:
        gdf = gpd.read_file(shp_path)
        if gdf.crs != "EPSG:4326":
            gdf = gdf.to_crs(epsg=4326)
        return gdf
    except Exception as e:
        st.error(f"Error memuat shapefile: {e}")
        return None

# Header dengan styling baru dan logo
try:
    logo = Image.open("Logo.png")
    col_logo1, col_header, col_logo2 = st.columns([1, 4, 1])
    
    with col_logo1:
        st.image(logo, width=100)
    
    with col_header:
        st.markdown('<p class="main-header">📊 Sistem Analisis Data Stunting</p>', unsafe_allow_html=True)
        st.markdown('<p class="sub-header">Dinas Kesehatan Kabupaten Kuningan</p>', unsafe_allow_html=True)
    
    with col_logo2:
        st.image(logo, width=100)
except:
    st.markdown('<p class="main-header">📊 Sistem Analisis Data Stunting</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-header">Dinas Kesehatan Kabupaten Kuningan</p>', unsafe_allow_html=True)

st.markdown("---")

# Sidebar dengan styling baru
with st.sidebar:
    st.markdown("### 🏥 DINKES KUNINGAN")
    st.markdown("---")
    st.markdown("### 📤 UPLOAD DATA")
    
    uploaded_file_gizi = st.file_uploader("📄 File Status Gizi", type=['xls', 'xlsx'], key='gizi')
    uploaded_file_sasaran = st.file_uploader("📄 File Sasaran Balita", type=['xls', 'xlsx'], key='sasaran')
    
    if uploaded_file_gizi and uploaded_file_sasaran:
        st.success("✅ File berhasil diupload!")
    
        # TAMBAHKAN CODE INI
        st.markdown("---")
        st.markdown("### 📅 INFORMASI DATA")
        
        # Pilih bulan data
        bulan_options = ['JANUARI', 'FEBRUARI', 'MARET', 'APRIL', 'MEI', 'JUNI',
                        'JULI', 'AGUSTUS', 'SEPTEMBER', 'OKTOBER', 'NOVEMBER', 'DESEMBER']
        pilih_bulan = st.selectbox("📊 Bulan Data Stunting:", bulan_options, key='bulan_data')
        
        # Input tanggal penarikan
        tanggal_penarikan = st.date_input(
            "📅 Tanggal Penarikan Data:",
            value=pd.to_datetime('today'),
            key='tanggal_penarikan'
        )
        
        # Format tanggal untuk ditampilkan
        tanggal_penarikan_str = tanggal_penarikan.strftime("%d %B %Y")

    st.markdown("---")
    st.markdown("### 📖 PANDUAN")
    with st.expander("💡 Cara Menggunakan"):
        st.markdown("""
        **1. Upload File**
        - Upload 2 file yang diperlukan
        - Format: .xls atau .xlsx
        
        **2. Tunggu Proses**
        - Sistem memproses otomatis
        - Progress ditampilkan
        
        **3. Lihat Hasil**
        - Jelajahi tab visualisasi
        - Interaksi dengan grafik
        
        **4. Download**
        - Unduh hasil analisis
        - Format CSV siap pakai
        - Download grafik dalam PNG
        """)
    
    with st.expander("📚 Tentang Indikator"):
        st.markdown("""
        **🔴 Stunting**
        Tinggi badan pendek untuk usia (TB/U)
        
        **🟡 Kurang Gizi**
        Berat badan kurang untuk usia (BB/U)
        
        **🟠 Wasting**
        Berat badan kurang untuk tinggi (BB/TB)
        """)

# Main content
if uploaded_file_gizi is None or uploaded_file_sasaran is None:
    st.info("👈 Silakan upload kedua file data di sidebar untuk memulai analisis.")
    
    # Metrics dengan styling baru
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("🏘️ Total Kecamatan", "32")
    with col2:
        st.metric("✅ Status Sistem", "Siap")
    with col3:
        st.metric("⚙️ Mode", "Proses Data & Visualisasi")
    with col4:
        st.metric("⏳ Status", "Menunggu")
    
    st.markdown("---")
    st.markdown("### 🎯 FITUR SISTEM")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.markdown("""
        <div class="feature-card">
            <h4>🗺️ Peta Interaktif</h4>
            <p>• Visualisasi shapefile geografis<br>
            • Sebaran stunting per desa<br>
            • Informasi detail saat hover<br>
            • Color-coded berdasarkan prevalensi</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="feature-card">
            <h4>📊 Analisis Komprehensif</h4>
            <p>• Grafik perbandingan antar wilayah<br>
            • Top ranking kecamatan<br>
            • Multi-indikator gizi<br>
            • Distribusi kategori status</p>
        </div>
        """, unsafe_allow_html=True)
    
    with col3:
        st.markdown("""
        <div class="feature-card">
            <h4>💾 Export & Download</h4>
            <p>• Download hasil ETL (Star Schema)<br>
            • Format CSV siap analisis<br>
            • Download grafik sebagai PNG<br>
            • Ringkasan statistik lengkap</p>
        </div>
        """, unsafe_allow_html=True)

else:
    # Simpan file temporary
    import tempfile
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as tmp_gizi:
        tmp_gizi.write(uploaded_file_gizi.getvalue())
        tmp_gizi_path = tmp_gizi.name
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as tmp_sasaran:
        tmp_sasaran.write(uploaded_file_sasaran.getvalue())
        tmp_sasaran_path = tmp_sasaran.name
    
    # Simpan info bulan dan tanggal ke session state
    if 'pilih_bulan' not in st.session_state:
        st.session_state.pilih_bulan = 'JANUARI'
    if 'tanggal_penarikan_str' not in st.session_state:
        st.session_state.tanggal_penarikan_str = pd.to_datetime('today').strftime("%d %B %Y")

    # Update dari input user jika ada
    if 'bulan_data' in locals():
        st.session_state.pilih_bulan = pilih_bulan
        st.session_state.tanggal_penarikan_str = tanggal_penarikan_str

    with st.spinner("🔄 Memproses data... Mohon tunggu..."):
        df_fact, df_wilayah, df_waktu, success, message = proses_etl(tmp_gizi_path, tmp_sasaran_path)

    # Hapus file temporary
    os.unlink(tmp_gizi_path)
    os.unlink(tmp_sasaran_path)
    
    if success:
        st.success(message)
        
        # Agregasi data per kecamatan
        df_agg = df_fact.groupby('puskesmas').agg({
            'jumlah_ditimbang_d': 'sum',
            'jumlah_stunting': 'sum',
            'jumlah_kurang_gizi': 'sum',
            'jumlah_wasting': 'sum',
            'sasaran_total': 'sum'
        }).reset_index()
        
        df_agg.columns = ['nama_kecamatan', 'jumlah_balita_ditimbang', 'jumlah_balita_stunting', 
                          'jumlah_balita_kurang_gizi', 'jumlah_balita_wasting', 'sasaran_total']
        
        df_agg['persentase_stunting'] = (df_agg['jumlah_balita_stunting'] / df_agg['jumlah_balita_ditimbang'] * 100).fillna(0)
        df_agg['persentase_kurang_gizi'] = (df_agg['jumlah_balita_kurang_gizi'] / df_agg['jumlah_balita_ditimbang'] * 100).fillna(0)
        df_agg['persentase_wasting'] = (df_agg['jumlah_balita_wasting'] / df_agg['jumlah_balita_ditimbang'] * 100).fillna(0)
        df_agg['persentase_sasaran'] = (df_agg['jumlah_balita_ditimbang'] / df_agg['sasaran_total'] * 100).fillna(0)
        
        # Ringkasan statistik dengan styling baru yang lebih informatif
        st.markdown("### 📈 RINGKASAN DATA STATISTIK STUNTING PER KABUPATEN KUNINGAN")
        
        total_ditimbang = int(df_agg['jumlah_balita_ditimbang'].sum())
        total_sasaran = int(df_agg['sasaran_total'].sum())
        total_stunting = int(df_agg['jumlah_balita_stunting'].sum())
        total_kurang_gizi = int(df_agg['jumlah_balita_kurang_gizi'].sum())
        total_wasting = int(df_agg['jumlah_balita_wasting'].sum())
        avg_stunting = df_agg['persentase_stunting'].mean()

        persen_ditimbang = (total_ditimbang / total_sasaran * 100) if total_sasaran > 0 else 0
        persen_stunting = (total_stunting / total_ditimbang * 100) if total_ditimbang > 0 else 0
        persen_kurang_gizi = (total_kurang_gizi / total_ditimbang * 100) if total_ditimbang > 0 else 0
        persen_wasting = (total_wasting / total_ditimbang * 100) if total_ditimbang > 0 else 0
        
        # Row dengan 4 kolom untuk card yang lebih besar
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 20px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); 
                        color: white; text-align: center;'>
                <div style='font-size: 14px; opacity: 0.9; margin-bottom: 8px;'>⚖️ BALITA DITIMBANG (D)</div>
                <div style='font-size: 32px; font-weight: 700; margin-bottom: 5px;'>{total_ditimbang:,}</div>
                <div style='font-size: 24px; font-weight: 600; color: #ffd700;'>{persen_ditimbang:.1f}% dari sasaran</div>
                <div style='font-size: 18px; opacity: 0.8; margin-top: 8px; border-top: 1px solid rgba(255,255,255,0.3); padding-top: 8px;'>
                    🎯 Sasaran: {total_sasaran:,} balita
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 20px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); 
                        color: white; text-align: center;'>
                <div style='font-size: 14px; opacity: 0.9; margin-bottom: 8px;'>📉 TOTAL STUNTING (S)</div>
                <div style='font-size: 32px; font-weight: 700; margin-bottom: 5px;'>{total_stunting:,}</div>
                <div style='font-size: 24px; font-weight: 600; color: #ffd700;'>{persen_stunting:.2f}% (S/D)</div>
                <div style='font-size: 18px; opacity: 0.8; margin-top: 8px; border-top: 1px solid rgba(255,255,255,0.3); padding-top: 8px;'>
                    dari {total_ditimbang:,} balita ditimbang
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 20px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); 
                        color: white; text-align: center;'>
                <div style='font-size: 14px; opacity: 0.9; margin-bottom: 8px;'>🍽️ UNDERWEIGHT (U)</div>
                <div style='font-size: 32px; font-weight: 700; margin-bottom: 5px;'>{total_kurang_gizi:,}</div>
                <div style='font-size: 24px; font-weight: 600; color: #ffd700;'>{persen_kurang_gizi:.2f}% (U/D)</div>
                <div style='font-size: 18px; opacity: 0.8; margin-top: 8px; border-top: 1px solid rgba(255,255,255,0.3); padding-top: 8px;'>
                    dari {total_ditimbang:,} balita ditimbang
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            st.markdown(f"""
            <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        padding: 20px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.1); 
                        color: white; text-align: center;'>
                <div style='font-size: 14px; opacity: 0.9; margin-bottom: 8px;'>⚠️ WASTING (W) </div>
                <div style='font-size: 32px; font-weight: 700; margin-bottom: 5px;'>{total_wasting:,}</div>
                <div style='font-size: 24px; font-weight: 600; color: #ffd700;'>{persen_wasting:.2f}% (W/D)</div>
                <div style='font-size: 18px; opacity: 0.8; margin-top: 8px; border-top: 1px solid rgba(255,255,255,0.3); padding-top: 8px;'>
                    dari {total_ditimbang:,} balita ditimbang
                </div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Tab untuk visualisasi dengan styling baru
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "🗺️ Peta Sebaran Stunting", 
            "📊 Perbandingan Stunting Antar Wilayah ", 
            "🎯 Sebaran Status Gizi Balita", 
            "📋 Tabel Data", 
            "💾 Download",
        ])
        
        with tab1:
            waktu_info = f"Bulan {st.session_state.pilih_bulan} (Penarikan: {st.session_state.tanggal_penarikan_str})"            
            st.markdown(
                f"### 🗺️ PETA SEBARAN STUNTING PER DESA DI KABUPATEN KUNINGAN "
                f"{pilih_bulan}"
            )
            
            # Load shapefile
            SHP_FILE_PATH = "data/ADMINISTRASIDESA_AR_25K.shp"
            data_gdf = load_shapefile(SHP_FILE_PATH)
            
            if data_gdf is not None:
                # Join data stunting dengan shapefile
                df_fact['desa_normalized'] = df_fact['desa'].str.strip().str.upper()
                data_gdf['NAMOBJ_normalized'] = data_gdf['NAMOBJ'].str.strip().str.upper()
                
                data_gdf_merged = data_gdf.merge(
                    df_fact[['desa_normalized', 'puskesmas', 'jumlah_ditimbang_d', 'sasaran_total', 
                             'persentase_ds', 'jumlah_stunting', 'persen_stunting']],
                    left_on='NAMOBJ_normalized',
                    right_on='desa_normalized',
                    how='left'
                )
                
                # Isi nilai NaN
                data_gdf_merged['jumlah_ditimbang_d'] = data_gdf_merged['jumlah_ditimbang_d'].fillna(0)
                data_gdf_merged['sasaran_total'] = data_gdf_merged['sasaran_total'].fillna(0)
                data_gdf_merged['persentase_ds'] = data_gdf_merged['persentase_ds'].fillna(0)
                data_gdf_merged['jumlah_stunting'] = data_gdf_merged['jumlah_stunting'].fillna(0)
                data_gdf_merged['persen_stunting'] = data_gdf_merged['persen_stunting'].fillna(0)
                data_gdf_merged['puskesmas'] = data_gdf_merged['puskesmas'].fillna('N/A')
                
                # ==================== FITUR PENCARIAN DESA ====================
                st.markdown("---")
                st.markdown("#### 🔍 Cari Desa")
                
                col_search1, col_search2 = st.columns([3, 1])
                
                with col_search1:
                    # Buat list desa untuk autocomplete
                    desa_list = sorted(data_gdf_merged['NAMOBJ'].dropna().unique().tolist())
                    
                    search_query = st.selectbox(
                        "Pilih atau ketik nama desa:",
                        options=[""] + desa_list,
                        index=0,
                        help="Ketik untuk mencari atau pilih dari dropdown"
                    )
                
                with col_search2:
                    clear_search = st.button("🔄 Reset", use_container_width=True)
                
                # Jika tombol reset diklik
                if clear_search:
                    search_query = ""
                    st.rerun()
                
                # ==================== END FITUR PENCARIAN ====================
                
                # Hitung bounds untuk zoom otomatis ke wilayah Kuningan
                bounds = data_gdf_merged.total_bounds  # [minx, miny, maxx, maxy]
                center_lat = (bounds[1] + bounds[3]) / 2
                center_lon = (bounds[0] + bounds[2]) / 2
                
                # Ambil jumlah kecamatan dan desa dari data ETL (df_fact)
                jumlah_kecamatan = 32
                jumlah_desa_dengan_data = 361
                
                # Layout: Peta di kiri (lebih besar), Legenda di kanan (lebih kecil)
                col_map, col_legend = st.columns([9.5, 3])
                
                with col_map:
                    # Buat peta Folium dengan tiles yang lebih bagus
                    m = folium.Map(
                        location=[center_lat, center_lon], 
                        zoom_start=11,
                        tiles='CartoDB positron',  # Tiles yang lebih bersih
                        control_scale=True,
                        zoom_control=True,
                        scrollWheelZoom=False,
                        dragging=True
                    )
                    
                    # Fungsi warna yang lebih detail
                    def get_color(persen_stunting):
                        if persen_stunting == 0:
                            return '#e0e0e0'
                        elif persen_stunting < 5:
                            return '#fff3cd'
                        elif persen_stunting < 10:
                            return '#ffcc80'
                        elif persen_stunting < 15:
                            return '#ff8c42'
                        elif persen_stunting < 20:
                            return '#ff6b6b'
                        else:
                            return '#d9534f'
                    
                    # Layer GeoJson dengan styling lebih baik
                    folium.GeoJson(
                        data_gdf_merged,
                        name="Stunting per Desa",
                        style_function=lambda feature: {
                            'fillColor': get_color(feature['properties'].get('persen_stunting', 0)),
                            'color': '#34495e',
                            'weight': 1.2,
                            'fillOpacity': 0.8,
                            'dashArray': '0'
                        },
                        highlight_function=lambda x: {
                            'fillColor': '#667eea',
                            'color': '#1a237e',
                            'weight': 3,
                            'fillOpacity': 0.9
                        },
                        tooltip=folium.GeoJsonTooltip(
                            fields=['NAMOBJ', 'WADMKC','puskesmas', 'jumlah_ditimbang_d', 'sasaran_total', 'persentase_ds', 
                                    'jumlah_stunting', 'persen_stunting'],
                            aliases=['🏘️ Desa:', '🏘️ Kecamatan','🏥 Puskesmas:', '⚖️ Ditimbang (D):', '🎯 Sasaran (S):', '📊 % Sasaran (D/S):', 
                                     '📉 Jml Stunting (JS):', '🔴 Prevalensi (JS/D):'],
                            localize=True,
                            sticky=False,
                            labels=True,
                            style="""
                                background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
                                border: 3px solid #667eea;
                                border-radius: 12px;
                                box-shadow: 0 6px 20px rgba(0,0,0,0.2);
                                padding: 12px 16px;
                                font-family: 'Poppins', sans-serif;
                                font-weight: 500;
                                font-size: 14px;
                                max-width: 300px;
                            """
                        )
                    ).add_to(m)
                    
                    # Tambahkan label kecamatan di peta
                    if 'WADMKC' in data_gdf_merged.columns:
                        # Agregasi per kecamatan untuk mendapatkan centroid
                        kecamatan_centroids = data_gdf_merged.groupby('WADMKC').apply(
                            lambda x: x.geometry.unary_union.centroid
                        ).reset_index()
                        kecamatan_centroids.columns = ['WADMKC', 'centroid']
                        
                        # Tambahkan marker untuk setiap kecamatan
                        for idx, row in kecamatan_centroids.iterrows():
                            folium.Marker(
                                location=[row['centroid'].y, row['centroid'].x],
                                icon=folium.DivIcon(html=f"""
                                    <div style="
                                        font-family: Arial, sans-serif;
                                        font-size: 10px;
                                        font-weight: 700;
                                        color: #2c3e50;
                                        text-shadow: 
                                            -1px -1px 0 rgba(255, 255, 255, 0.9),
                                            1px -1px 0 rgba(255, 255, 255, 0.9),
                                            -1px 1px 0 rgba(255, 255, 255, 0.9),
                                            1px 1px 0 rgba(255, 255, 255, 0.9),
                                            0 0 3px rgba(255, 255, 255, 0.7);
                                        white-space: nowrap;
                                        text-align: center;
                                        text-transform: uppercase;
                                        letter-spacing: 0.5px;
                                        transform: translateX(-50%);
                                        margin-left: 50%;
                                    ">
                                        {row['WADMKC']}
                                    </div>
                                """)
                            ).add_to(m)
                    
                    # Jika ada pencarian desa, tambahkan marker
                    if search_query:
                        search_result = data_gdf_merged[data_gdf_merged['NAMOBJ'] == search_query]
                        
                        if not search_result.empty:
                            result = search_result.iloc[0]
                            # Ambil centroid dari geometry desa
                            centroid = result.geometry.centroid
                            
                            folium.Marker(
                                location=[centroid.y, centroid.x],
                                popup=folium.Popup(f"""
                                    <div style='width: 200px; font-family: Poppins;'>
                                        <h4 style='color: #667eea; margin: 0 0 10px 0;'>📍 {result['NAMOBJ']}</h4>
                                        <b>Prevalensi:</b> {result['persen_stunting']:.2f}%<br>
                                        <b>Stunting:</b> {int(result['jumlah_stunting'])} balita<br>
                                        <b>Puskesmas:</b> {result['puskesmas']}
                                    </div>
                                """, max_width=250),
                                icon=folium.Icon(color='red', icon='info-sign', prefix='glyphicon'),
                                tooltip=f"📍 {result['NAMOBJ']}"
                            ).add_to(m)
                            
                            # Zoom ke desa yang dicari
                            m.fit_bounds([[centroid.y - 0.02, centroid.x - 0.02], 
                                          [centroid.y + 0.02, centroid.x + 0.02]])
                        else:
                            # Fit bounds agar hanya menampilkan wilayah Kuningan
                            m.fit_bounds([[bounds[1], bounds[0]], [bounds[3], bounds[2]]])
                    else:
                        # Fit bounds agar hanya menampilkan wilayah Kuningan
                        m.fit_bounds([[bounds[1], bounds[0]], [bounds[3], bounds[2]]])
                    
                    # Legend prevalensi stunting
                    legend_html = '''
                    <div style="position: fixed; 
                                bottom: 50px; left: 50px; width: 220px; 
                                background: linear-gradient(145deg, #ffffff 0%, #f8f9fa 100%); 
                                border: 3px solid #667eea; 
                                border-radius: 16px;
                                z-index: 9999; 
                                padding: 18px;
                                box-shadow: 0 8px 25px rgba(0,0,0,0.2);
                                font-family: 'Poppins', sans-serif;">

                    <p style="margin: 0 0 12px 0; font-weight: 700; font-size: 16px; color: #667eea; text-align: center;">
                    📊 Prevalensi Stunting</p>

                    <p style="margin: 6px 0;">
                    <i style="background:#e0e0e0; width: 30px; height: 14px; 
                    display: inline-block; border-radius: 4px; margin-right: 10px; border: 1px solid #ccc;"></i>
                    <span style="font-size: 13px; font-weight: 500;">Tidak ada data</span>
                    </p>

                    <p style="margin: 6px 0;">
                    <i style="background:#fff3cd; width: 30px; height: 14px; 
                    display: inline-block; border-radius: 4px; margin-right: 10px; border: 1px solid #ffeeba;"></i>
                    <span style="font-size: 13px; font-weight: 500;">&lt; 5% (Sangat Rendah)</span>
                    </p>

                    <p style="margin: 6px 0;">
                    <i style="background:#ffcc80; width: 30px; height: 14px; 
                    display: inline-block; border-radius: 4px; margin-right: 10px; border: 1px solid #ffb84d;"></i>
                    <span style="font-size: 13px; font-weight: 500;">5–15% (Sedang)</span>
                    </p>

                    <p style="margin: 6px 0;">
                    <i style="background:#ff6b6b; width: 30px; height: 14px; 
                    display: inline-block; border-radius: 4px; margin-right: 10px; border: 1px solid #ff5252;"></i>
                    <span style="font-size: 13px; font-weight: 500;">15–20% (Tinggi)</span>
                    </p>

                    <p style="margin: 6px 0;">
                    <i style="background:#d9534f; width: 30px; height: 14px; 
                    display: inline-block; border-radius: 4px; margin-right: 10px; border: 1px solid #c9302c;"></i>
                    <span style="font-size: 13px; font-weight: 500;">&gt; 20% (Sangat Tinggi)</span>
                    </p>

                    </div>
                    '''
                    m.get_root().html.add_child(folium.Element(legend_html))
                    
                    # Tampilkan peta dengan ukuran lebih besar
                    st_folium(m, width=1200, height=800, returned_objects=[])
                
                with col_legend:
                    # Informasi Kabupaten Kuningan
                    st.markdown("""
                    <div style='background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                                padding: 20px; border-radius: 12px; color: white; 
                                box-shadow: 0 4px 15px rgba(0,0,0,0.2);'>
                        <h3 style='margin: 0; text-align: center; color: white; font-size: 18px;'>
                        📍 Kabupaten Kuningan</h3>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    # Batas Wilayah
                    st.markdown("""
                    <div style='background: white; padding: 15px; border-radius: 10px; 
                                border: 2px solid #667eea; margin-bottom: 15px;'>
                        <h4 style='color: #667eea; margin: 0 0 10px 0; font-size: 15px;'>🗺️ Batas Wilayah</h4>
                        <p style='margin: 5px 0; font-size: 12px;'>
                            <b>Utara:</b><br>Kab. Cirebon & Majalengka
                        </p>
                        <p style='margin: 5px 0; font-size: 12px;'>
                            <b>Timur:</b><br>Kab. Brebes
                        </p>
                        <p style='margin: 5px 0; font-size: 12px;'>
                            <b>Selatan:</b><br>Kab. Cilacap
                        </p>
                        <p style='margin: 5px 0; font-size: 12px;'>
                            <b>Barat:</b><br>Kab. Majalengka
                        </p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    # Statistik Wilayah
                    st.markdown(f"""
                    <div style='background: white; padding: 15px; border-radius: 10px; 
                                border: 2px solid #667eea;'>
                        <h4 style='color: #667eea; margin: 0 0 10px 0; font-size: 15px;'>📊 Statistik Wilayah</h4>
                        <div style='background: #f8f9fa; padding: 10px; border-radius: 8px; margin: 8px 0;'>
                            <p style='margin: 0; font-size: 12px;'><b>Jumlah Kecamatan:</b></p>
                            <p style='margin: 5px 0 0 0; font-size: 22px; font-weight: bold; color: #667eea;'>
                                {jumlah_kecamatan}
                            </p>
                        </div>
                        <div style='background: #f8f9fa; padding: 10px; border-radius: 8px; margin: 8px 0;'>
                            <p style='margin: 0; font-size: 12px;'><b>Jumlah Desa:<br></b></p>
                            <p style='margin: 5px 0 0 0; font-size: 22px; font-weight: bold; color: #667eea;'>
                                {jumlah_desa_dengan_data}
                            </p>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("---")
                
                # Tombol Download Peta
                st.markdown("#### 💾 Download Peta")
                with st.spinner("🔄 Membuat peta statis untuk download..."):
                    map_img_bytes = create_static_map_image(
                        data_gdf_merged, 
                        "PETA SEBARAN STUNTING PER DESA - KABUPATEN KUNINGAN BULAN " f"{pilih_bulan}" 
                    )
                    if map_img_bytes:
                        create_download_button_for_map(map_img_bytes, "peta_sebaran_stunting_kuningan")
                        st.info("💡 Peta yang didownload adalah versi statis dengan resolusi tinggi (300 DPI) yang mencakup label nama desa dan persentase stunting.")
                
                st.markdown("---")
                
                st.markdown("""
                <div class="info-box">
                    <b>💡 Cara Membaca Peta</b><br><br>
                    🎨 <b>Warna wilayah</b> menunjukkan tingkat prevalensi stunting (semakin gelap merah, semakin tinggi prevalensi)<br><br>
                    🖱️ <b>Klik pada wilayah desa</b> untuk melihat informasi detail:<br>
                    &nbsp;&nbsp;&nbsp;&nbsp;• Nama Desa & Kecamatan<br>
                    &nbsp;&nbsp;&nbsp;&nbsp;• Nama Puskesmas<br>
                    &nbsp;&nbsp;&nbsp;&nbsp;• Jumlah Balita Ditimbang & Sasaran<br>
                    &nbsp;&nbsp;&nbsp;&nbsp;• Persentase Pencapaian Sasaran<br>
                    &nbsp;&nbsp;&nbsp;&nbsp;• Jumlah & Prevalensi Stunting<br><br>
                    🔍 <b>Gunakan scroll/zoom</b> untuk melihat detail wilayah tertentu<br><br>
                    🔎 <b>Gunakan fitur pencarian di atas</b> untuk mencari desa tertentu dan melihat lokasinya di peta<br><br>
                    🏘️ <b>Label kecamatan</b> ditampilkan langsung di peta untuk memudahkan identifikasi wilayah
                </div>
                """, unsafe_allow_html=True)
                
                st.markdown("---")
                                
                # Row 3: Top Desa, Kecamatan, dan Puskesmas
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.markdown("#### 🔴 10 Desa dengan Stunting Tertinggi")
                    top_desa = data_gdf_merged[data_gdf_merged['persen_stunting'] > 0].nlargest(10, 'persen_stunting')
                    
                    for idx, row in top_desa.iterrows():
                        with st.container():
                            st.markdown(f"""
                            <div style='background: linear-gradient(135deg, #fff5f5 0%, #ffe0e0 100%); 
                                        padding: 10px; border-radius: 8px; margin: 5px 0; 
                                        border-left: 4px solid #d9534f;'>
                                <b style='color: #d9534f;'>{row['NAMOBJ']}</b> 
                                <span style='color: #666;'>(Puskesmas {row['puskesmas']})</span><br>
                                <span style='font-size: 18px; font-weight: 700; color: #d9534f;'>{row['persen_stunting']:.2f}%</span> 
                                <span style='color: #666;'>• {int(row['jumlah_stunting'])} dari {int(row['jumlah_ditimbang_d'])} balita</span>
                            </div>
                            """, unsafe_allow_html=True)
                
                with col2:
                    st.markdown("#### 🔴 10 Kecamatan dengan Stunting Tertinggi")
                    
                    # Agregasi data per kecamatan
                    if 'WADMKC' in data_gdf_merged.columns:
                        kecamatan_col = 'WADMKC'
                    else:
                        kecamatan_col = 'puskesmas'
                    
                    kecamatan_agg = data_gdf_merged.groupby(kecamatan_col).agg({
                        'jumlah_stunting': 'sum',
                        'jumlah_ditimbang_d': 'sum'
                    }).reset_index()
                    
                    kecamatan_agg['persen_stunting'] = (kecamatan_agg['jumlah_stunting'] / kecamatan_agg['jumlah_ditimbang_d'] * 100).fillna(0)
                    top_kecamatan = kecamatan_agg[kecamatan_agg['persen_stunting'] > 0].nlargest(10, 'persen_stunting')
                    
                    for idx, row in top_kecamatan.iterrows():
                        with st.container():
                            st.markdown(f"""
                            <div style='background: linear-gradient(135deg, #fff5f5 0%, #ffe0e0 100%); 
                                        padding: 10px; border-radius: 8px; margin: 5px 0; 
                                        border-left: 4px solid #d9534f;'>
                                <b style='color: #d9534f;'>{row[kecamatan_col]}</b><br>
                                <span style='font-size: 18px; font-weight: 700; color: #d9534f;'>{row['persen_stunting']:.2f}%</span> 
                                <span style='color: #666;'>• {int(row['jumlah_stunting'])} dari {int(row['jumlah_ditimbang_d'])} balita</span>
                            </div>
                            """, unsafe_allow_html=True)
                
                with col3:
                    st.markdown("#### 🔴 10 Puskesmas dengan Stunting Tertinggi")
                    
                    # Agregasi data per puskesmas
                    puskesmas_agg = data_gdf_merged.groupby('puskesmas').agg({
                        'jumlah_stunting': 'sum',
                        'jumlah_ditimbang_d': 'sum'
                    }).reset_index()
                    
                    puskesmas_agg['persen_stunting'] = (puskesmas_agg['jumlah_stunting'] / puskesmas_agg['jumlah_ditimbang_d'] * 100).fillna(0)
                    top_puskesmas = puskesmas_agg[puskesmas_agg['persen_stunting'] > 0].nlargest(10, 'persen_stunting')
                    
                    for idx, row in top_puskesmas.iterrows():
                        with st.container():
                            st.markdown(f"""
                            <div style='background: linear-gradient(135deg, #fff5f5 0%, #ffe0e0 100%); 
                                        padding: 10px; border-radius: 8px; margin: 5px 0; 
                                        border-left: 4px solid #d9534f;'>
                                <b style='color: #d9534f;'>{row['puskesmas']}</b><br>
                                <span style='font-size: 18px; font-weight: 700; color: #d9534f;'>{row['persen_stunting']:.2f}%</span> 
                                <span style='color: #666;'>• {int(row['jumlah_stunting'])} dari {int(row['jumlah_ditimbang_d'])} balita</span>
                            </div>
                            """, unsafe_allow_html=True)
                                            
            else:
                st.error("⚠️ File shapefile tidak ditemukan di folder 'data/'.")
                st.info("📁 Pastikan file shapefile tersedia di folder 'data/ADMINISTRASIDESA_AR_25K.shp'")
        
        with tab2:
            waktu_info = f"Bulan {st.session_state.pilih_bulan} (Penarikan: {st.session_state.tanggal_penarikan_str})"            
            st.markdown(f"### 📊 PERBANDINGAN ANTAR WILAYAH BULAN "f"{pilih_bulan}")
            
            # Filter untuk memilih level perbandingan
            level_perbandingan = st.selectbox(
                "📍 Tampilkan Data:",
                ["Puskesmas", "Kecamatan", "Desa"],
                key="level_perbandingan"
            )
            
            # Tentukan dataframe berdasarkan pilihan
            if level_perbandingan == "Puskesmas":
                df_display_source = df_agg.copy()
                nama_kolom = 'nama_kecamatan'
                jumlah_max = len(df_agg)
                jumlah_default = min(15, jumlah_max)
            elif level_perbandingan == "Kecamatan":
                # Load shapefile untuk mendapatkan data kecamatan
                SHP_FILE_PATH = "data/ADMINISTRASIDESA_AR_25K.shp"
                data_gdf = load_shapefile(SHP_FILE_PATH)
                
                if data_gdf is not None:
                    # Merge df_fact dengan shapefile untuk mendapatkan kecamatan
                    df_fact['desa_normalized'] = df_fact['desa'].str.strip().str.upper()
                    data_gdf['NAMOBJ_normalized'] = data_gdf['NAMOBJ'].str.strip().str.upper()
                    
                    df_with_kec = df_fact.merge(
                        data_gdf[['NAMOBJ_normalized', 'WADMKC']],
                        left_on='desa_normalized',
                        right_on='NAMOBJ_normalized',
                        how='left'
                    )
                    
                    # Agregasi per kecamatan
                    df_kec_agg = df_with_kec.groupby('WADMKC').agg({
                        'jumlah_ditimbang_d': 'sum',
                        'jumlah_stunting': 'sum'
                    }).reset_index()
                    
                    df_kec_agg['persentase_stunting'] = (df_kec_agg['jumlah_stunting'] / df_kec_agg['jumlah_ditimbang_d'] * 100).fillna(0)
                    df_kec_agg.columns = ['nama_kecamatan', 'jumlah_balita_ditimbang', 'jumlah_balita_stunting', 'persentase_stunting']
                    
                    df_display_source = df_kec_agg.copy()
                    nama_kolom = 'nama_kecamatan'
                    jumlah_max = len(df_kec_agg)
                    jumlah_default = min(15, jumlah_max)
                else:
                    st.error("⚠️ Shapefile tidak ditemukan. Tidak dapat menampilkan data per kecamatan.")
                    df_display_source = pd.DataFrame()
                    nama_kolom = 'nama_kecamatan'
                    jumlah_max = 0
                    jumlah_default = 0
            else:  # Desa
                # Agregasi untuk desa (karena df_fact sudah punya data per desa)
                df_display_source = df_fact[['desa', 'jumlah_ditimbang_d', 'jumlah_stunting', 'persen_stunting']].copy()
                df_display_source.columns = ['nama_desa', 'jumlah_balita_ditimbang', 'jumlah_balita_stunting', 'persentase_stunting']
                
                nama_kolom = 'nama_desa'
                jumlah_max = len(df_display_source)
                jumlah_default = min(15, jumlah_max)
            
            if not df_display_source.empty:
                col1, col2 = st.columns([2, 1])
                with col1:
                    jumlah_tampil = st.slider(
                        "🔢 Jumlah yang ditampilkan:", 
                        min_value=5, 
                        max_value=max(jumlah_max, 5), 
                        value=jumlah_default,
                        key="jumlah_slider"
                    )
                with col2:
                    urutan = st.radio("📈 Urutan:", ["Tertinggi", "Terendah"], key="urutan_radio")
                
                st.markdown(f"#### 📊 {level_perbandingan} dengan Kasus Stunting Tertinggi Bulan " f"{pilih_bulan}")
                
                # Sorting berdasarkan urutan
                if urutan == "Tertinggi":
                    df_display = df_display_source.nlargest(jumlah_tampil, 'persentase_stunting')
                else:
                    df_display = df_display_source.nsmallest(jumlah_tampil, 'persentase_stunting')
                
                # Membuat grafik
                fig_bar = go.Figure()
                
                fig_bar.add_trace(go.Bar(
                    y=df_display[nama_kolom],
                    x=df_display['persentase_stunting'],
                    orientation='h',
                    text=[f"{persen:.1f}% ({int(jml)} balita)" 
                        for persen, jml in zip(df_display['persentase_stunting'], df_display['jumlah_balita_stunting'])],
                    textposition='outside',
                    marker=dict(
                        color=df_display['persentase_stunting'],
                        colorscale=[[0, '#fff3cd'], [0.5, '#ff8c42'], [1, '#d9534f']],
                        showscale=True,
                        colorbar=dict(
                            title=dict(
                                text="Persentase (%)",
                                font=dict(size=12, family='Poppins')
                            ),
                            tickfont=dict(size=11, family='Poppins')
                        )
                    ),
                    hovertemplate='<b>%{y}</b><br>Persentase: %{x:.2f}%<br><extra></extra>'
                ))
                
                fig_bar.update_layout(
                    height=max(450, jumlah_tampil * 35),
                    xaxis_title='Persentase Stunting (%)',
                    yaxis_title='',
                    yaxis={'categoryorder':'total ascending' if urutan == "Tertinggi" else 'total descending'},
                    font=dict(size=12, family='Poppins'),
                    margin=dict(l=150, r=150, t=30, b=50),
                    plot_bgcolor='rgba(0,0,0,0)',
                    paper_bgcolor='rgba(0,0,0,0)'
                )
                st.plotly_chart(fig_bar, use_container_width=True, config={'displayModeBar': False})
                
                # Tombol download grafik
                create_download_button_for_chart(
                    fig_bar, 
                    f"top_{level_perbandingan.lower()}_stunting_{urutan.lower()}",
                    f"Top {jumlah_tampil} {level_perbandingan} dengan Stunting {urutan} {waktu_info}"
                )
                             
        with tab3:
            waktu_info = f"Bulan {st.session_state.pilih_bulan} (Penarikan: {st.session_state.tanggal_penarikan_str})"            
            st.markdown(f"### 🎯 SEBARAN STATUS GIZI BALITA KABUPATEN KUNINGAN")
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                st.markdown(f"#### 📊 Komposisi Status Gizi Balita Bulan " f"{pilih_bulan}")
                
                total_normal = total_ditimbang - total_stunting - total_kurang_gizi - total_wasting
                
                labels = ['Stunting', 'Underweight', 'Wasting', 'Normal']
                values = [total_stunting, total_kurang_gizi, total_wasting, total_normal]
                colors = ['#d9534f', '#f0ad4e', '#ff8c42', '#5bc0de']
                
                fig_pie = go.Figure(data=[go.Pie(
                    labels=labels,
                    values=values,
                    hole=0.5,
                    marker_colors=colors,
                    textinfo='label+percent',
                    textfont=dict(size=13, family='Poppins', color='white'),
                    hovertemplate='<b>%{label}</b><br>Jumlah: %{value:,} balita<br>Persentase: %{percent}<extra></extra>'
                )])
                
                fig_pie.update_layout(
                    height=500,
                    title_text="",
                    title_font=dict(size=16, family='Poppins', color='#667eea'),
                    font=dict(size=12, family='Poppins'),
                    showlegend=True,
                    legend=dict(
                        orientation="v",
                        yanchor="middle",
                        y=0.5,
                        xanchor="left",
                        x=1.02
                    ),
                    margin=dict(t=80, b=20, l=20, r=150),
                    paper_bgcolor='rgba(0,0,0,0)'
                )
                st.plotly_chart(fig_pie, use_container_width=True, config={'displayModeBar': False})
                
                # Tombol download grafik
                create_download_button_for_chart(
                    fig_pie, 
                    "komposisi_status_gizi_balita",
                    f"Komposisi Status Gizi Balita Bulan " f"{pilih_bulan}"
                )
            
            with col2:
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown("<br>", unsafe_allow_html=True)

                
                # Row 1: Stunting dan Underweight
                col_card1, col_card2 = st.columns(2, gap="medium")
                
                with col_card1:
                    # Card Stunting
                    persen_stunting_total = (total_stunting / total_ditimbang * 100) if total_ditimbang > 0 else 0
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #fff5f5 0%, #ffe0e0 100%); 
                                padding: 12px 15px; border-radius: 10px; 
                                border-left: 4px solid #d9534f; box-shadow: 0 2px 8px rgba(0,0,0,0.08);'>
                        <h4 style='color: #d9534f; margin: 0 0 8px 0; font-size: 14px;'>
                            🔴 Stunting (S)
                        </h4>
                        <div style='display: flex; justify-content: space-between; align-items: center;'>
                            <div>
                                <p style='margin: 0; font-size: 10px; color: #888;'>Jumlah</p>
                                <p style='margin: 2px 0 0 0; font-size: 20px; font-weight: 700; color: #d9534f;'>
                                    {total_stunting:,}
                                </p>
                            </div>
                            <div style='text-align: right;'>
                                <p style='margin: 0; font-size: 10px; color: #888;'>Ditimbang (D)</p>
                                <p style='margin: 2px 0 0 0; font-size: 16px; font-weight: 600; color: #666;'>
                                    {total_ditimbang:,}
                                </p>
                            </div>
                        </div>
                        <div style='margin-top: 8px; padding-top: 8px; border-top: 1px solid #ffcccc;'>
                            <p style='margin: 0; font-size: 16px; font-weight: 700; color: #d9534f; text-align: center;'>
                                {persen_stunting_total:.2f}% (S/D)
                            </p>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col_card2:
                    # Card Underweight (Kurang Gizi)
                    persen_kurang_gizi_total = (total_kurang_gizi / total_ditimbang * 100) if total_ditimbang > 0 else 0
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #fff9f0 0%, #ffe8cc 100%); 
                                padding: 12px 15px; border-radius: 10px; 
                                border-left: 4px solid #f0ad4e; box-shadow: 0 2px 8px rgba(0,0,0,0.08);'>
                        <h4 style='color: #f0ad4e; margin: 0 0 8px 0; font-size: 14px;'>
                            🟡 Underweight (U)
                        </h4>
                        <div style='display: flex; justify-content: space-between; align-items: center;'>
                            <div>
                                <p style='margin: 0; font-size: 10px; color: #888;'>Jumlah</p>
                                <p style='margin: 2px 0 0 0; font-size: 20px; font-weight: 700; color: #f0ad4e;'>
                                    {total_kurang_gizi:,}
                                </p>
                            </div>
                            <div style='text-align: right;'>
                                <p style='margin: 0; font-size: 10px; color: #888;'>Ditimbang (D)</p>
                                <p style='margin: 2px 0 0 0; font-size: 16px; font-weight: 600; color: #666;'>
                                    {total_ditimbang:,}
                                </p>
                            </div>
                        </div>
                        <div style='margin-top: 8px; padding-top: 8px; border-top: 1px solid #ffe0b3;'>
                            <p style='margin: 0; font-size: 16px; font-weight: 700; color: #f0ad4e; text-align: center;'>
                                {persen_kurang_gizi_total:.2f}% (U/D)
                            </p>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                st.markdown("<br>", unsafe_allow_html=True)
                
                # Row 2: Wasting dan Normal
                col_card3, col_card4 = st.columns(2, gap="medium")
                
                with col_card3:
                    # Card Wasting
                    persen_wasting_total = (total_wasting / total_ditimbang * 100) if total_ditimbang > 0 else 0
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #fff5f0 0%, #ffd9cc 100%); 
                                padding: 12px 15px; border-radius: 10px; 
                                border-left: 4px solid #ff8c42; box-shadow: 0 2px 8px rgba(0,0,0,0.08);'>
                        <h4 style='color: #ff8c42; margin: 0 0 8px 0; font-size: 14px;'>
                            🟠 Wasting (W)
                        </h4>
                        <div style='display: flex; justify-content: space-between; align-items: center;'>
                            <div>
                                <p style='margin: 0; font-size: 10px; color: #888;'>Jumlah</p>
                                <p style='margin: 2px 0 0 0; font-size: 20px; font-weight: 700; color: #ff8c42;'>
                                    {total_wasting:,}
                                </p>
                            </div>
                            <div style='text-align: right;'>
                                <p style='margin: 0; font-size: 10px; color: #888;'>Ditimbang (D)</p>
                                <p style='margin: 2px 0 0 0; font-size: 16px; font-weight: 600; color: #666;'>
                                    {total_ditimbang:,}
                                </p>
                            </div>
                        </div>
                        <div style='margin-top: 8px; padding-top: 8px; border-top: 1px solid #ffccb3;'>
                            <p style='margin: 0; font-size: 16px; font-weight: 700; color: #ff8c42; text-align: center;'>
                                {persen_wasting_total:.2f}% (W/D)
                            </p>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col_card4:
                    # Card Normal
                    persen_normal_total = (total_normal / total_ditimbang * 100) if total_ditimbang > 0 else 0
                    st.markdown(f"""
                    <div style='background: linear-gradient(135deg, #f0f8ff 0%, #d9eeff 100%); 
                                padding: 12px 15px; border-radius: 10px; 
                                border-left: 4px solid #5bc0de; box-shadow: 0 2px 8px rgba(0,0,0,0.08);'>
                        <h4 style='color: #5bc0de; margin: 0 0 8px 0; font-size: 14px;'>
                            🔵 Normal (N)
                        </h4>
                        <div style='display: flex; justify-content: space-between; align-items: center;'>
                            <div>
                                <p style='margin: 0; font-size: 10px; color: #888;'>Jumlah</p>
                                <p style='margin: 2px 0 0 0; font-size: 20px; font-weight: 700; color: #5bc0de;'>
                                    {total_normal:,}
                                </p>
                            </div>
                            <div style='text-align: right;'>
                                <p style='margin: 0; font-size: 10px; color: #888;'>Ditimbang (D)</p>
                                <p style='margin: 2px 0 0 0; font-size: 16px; font-weight: 600; color: #666;'>
                                    {total_ditimbang:,}
                                </p>
                            </div>
                        </div>
                        <div style='margin-top: 8px; padding-top: 8px; border-top: 1px solid #b3e0ff;'>
                            <p style='margin: 0; font-size: 16px; font-weight: 700; color: #5bc0de; text-align: center;'>
                                {persen_normal_total:.2f}% (N/D)
                            </p>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)            
            
            st.markdown("""
            <div class="info-box">
                <b>📚 Penjelasan Indikator Gizi</b><br><br>
                🔴 <b>BB/U (Berat Badan per Usia)</b><br>
                &nbsp;&nbsp;&nbsp;&nbsp;Mengukur kecukupan berat badan anak sesuai usianya<br><br>
                🟡 <b>TB/U (Tinggi Badan per Usia)</b><br>
                &nbsp;&nbsp;&nbsp;&nbsp;Mengukur stunting atau kekurangan gizi kronis<br><br>
                🟠 <b>BB/TB (Berat Badan per Tinggi Badan)</b><br>
                &nbsp;&nbsp;&nbsp;&nbsp;Mengukur wasting atau kekurangan gizi akut
            </div>
            """, unsafe_allow_html=True)
        
        with tab4:
            waktu_info = f"Bulan {st.session_state.pilih_bulan} (Penarikan: {st.session_state.tanggal_penarikan_str})"            
            st.markdown(f"### 📋 DATA STUNTING PER WILAYAH KABUPATEN KUNINGAN DALAM TABLE BULAN "f"{pilih_bulan}")
            
            df_desa_table = df_fact.copy()

            
            # Load shapefile untuk mendapatkan data kecamatan
            SHP_FILE_PATH = "data/ADMINISTRASIDESA_AR_25K.shp"
            data_gdf = load_shapefile(SHP_FILE_PATH)
            
            if data_gdf is not None:
                # Buat mapping desa ke kecamatan dari shapefile
                desa_kecamatan_map = data_gdf[['NAMOBJ', 'WADMKC']].copy()
                desa_kecamatan_map['NAMOBJ'] = desa_kecamatan_map['NAMOBJ'].str.strip().str.upper()
                desa_kecamatan_map = desa_kecamatan_map.drop_duplicates(subset=['NAMOBJ'])
                desa_kecamatan_map = dict(zip(desa_kecamatan_map['NAMOBJ'], desa_kecamatan_map['WADMKC']))
            else:
                desa_kecamatan_map = {}
            
            # Siapkan data desa dengan kolom yang sesuai
            df_display = df_desa_table[['desa', 'puskesmas', 'sasaran_total', 'jumlah_ditimbang_d', 
                                        'persentase_ds', 'jumlah_stunting', 'persen_stunting',
                                        'jumlah_kurang_gizi', 'persen_kurang_gizi',
                                        'jumlah_wasting', 'persen_wasting']].copy()
            
            df_display.columns = ['nama_desa', 'nama_puskesmas', 'sasaran_total', 'jumlah_balita_ditimbang',
                                'persentase_sasaran', 'jumlah_balita_stunting', 'persentase_stunting',
                                'jumlah_balita_kurang_gizi', 'persentase_kurang_gizi',
                                'jumlah_balita_wasting', 'persentase_wasting']
            
            # Tambahkan kolom kecamatan dari mapping shapefile
            df_display['nama_kecamatan'] = df_display['nama_desa'].map(desa_kecamatan_map)
            df_display['nama_kecamatan'] = df_display['nama_kecamatan'].fillna('N/A')
            
            # Tambahkan kategori untuk desa
            df_display['kategori'] = pd.cut(
                df_display['persentase_stunting'],
                bins=[0, 5, 10, 20, 100],
                labels=['Rendah (<5%)', 'Sedang (5-10%)', 'Tinggi (10-20%)', 'Sangat Tinggi (>20%)']
            )
            
            st.markdown("---")
            
            col1, col2 = st.columns([3, 1])
            with col1:
                search_term = st.text_input(
                    "🔍 Cari Desa:", 
                    placeholder="Ketik nama desa...",
                    key="search_wilayah"
                )
            with col2:
                sort_by = st.selectbox(
                    "📊 Urutkan:", 
                    ["Nama", "% Stunting", "Jml Stunting", "Jml Ditimbang"],
                    key="sort_by_table"
                )
            
            # Filter berdasarkan pencarian
            if search_term:
                df_display = df_display[df_display['nama_desa'].str.contains(search_term.upper(), na=False)]
            
            # Sorting
            if sort_by == "Nama":
                df_display = df_display.sort_values('nama_desa')
            elif sort_by == "% Stunting":
                df_display = df_display.sort_values('persentase_stunting', ascending=False)
            elif sort_by == "Jml Stunting":
                df_display = df_display.sort_values('jumlah_balita_stunting', ascending=False)
            else:
                df_display = df_display.sort_values('jumlah_balita_ditimbang', ascending=False)
            
            # Siapkan tabel untuk ditampilkan dengan urutan: Desa, Kecamatan, Puskesmas
            df_table = df_display[['nama_desa', 'nama_kecamatan', 'nama_puskesmas', 'sasaran_total', 'jumlah_balita_ditimbang', 
                                'persentase_sasaran', 'jumlah_balita_stunting', 
                                'persentase_stunting', 'jumlah_balita_kurang_gizi', 
                                'persentase_kurang_gizi', 'jumlah_balita_wasting', 
                                'persentase_wasting', 'kategori']].copy()
            
            df_table.columns = ['Desa', 'Kecamatan', 'Puskesmas', 'Sasaran (Sa)', 'Ditimbang (D)', 'Prevalensi Sasaran (D/Sa)', 'Stunting (S)', 'Prevalensi Stunting (S/D)', 
                            'Underweight (U)', 'Prevalensi Underweight (U/D)', 'Wasting (W)', 'Prevalensi Wasting (W/D)', 'Kategori']
            
            # Format angka
            df_table['Sasaran (Sa)'] = df_table['Sasaran (Sa)'].apply(lambda x: f"{int(x):,}")
            df_table['Ditimbang (D)'] = df_table['Ditimbang (D)'].apply(lambda x: f"{int(x):,}")
            df_table['Prevalensi Sasaran (D/Sa)'] = df_table['Prevalensi Sasaran (D/Sa)'].apply(lambda x: f"{x:.2f}%")
            df_table['Stunting (S)'] = df_table['Stunting (S)'].apply(lambda x: f"{int(x):,}")
            df_table['Prevalensi Stunting (S/D)'] = df_table['Prevalensi Stunting (S/D)'].apply(lambda x: f"{x:.2f}%")
            df_table['Underweight (U)'] = df_table['Underweight (U)'].apply(lambda x: f"{int(x):,}")
            df_table['Prevalensi Underweight (U/D)'] = df_table['Prevalensi Underweight (U/D)'].apply(lambda x: f"{x:.2f}%")
            df_table['Wasting (W)'] = df_table['Wasting (W)'].apply(lambda x: f"{int(x):,}")
            df_table['Prevalensi Wasting (W/D)'] = df_table['Prevalensi Wasting (W/D)'].apply(lambda x: f"{x:.2f}%")
            
            # Fungsi highlight berdasarkan kategori
            def highlight_kategori(row):
                if 'Sangat Tinggi' in str(row['Kategori']):
                    return ['background-color: #ffcccc'] * len(row)
                elif 'Tinggi' in str(row['Kategori']):
                    return ['background-color: #ffe6cc'] * len(row)
                elif 'Sedang' in str(row['Kategori']):
                    return ['background-color: #fff4cc'] * len(row)
                else:
                    return ['background-color: #d4edda'] * len(row)
            
            df_styled = df_table.style.apply(highlight_kategori, axis=1)
            
            st.dataframe(df_styled, use_container_width=True, height=500)
            
            # Info jumlah data
            total_data = len(df_fact)
            st.info(f"📊 Menampilkan **{len(df_display)}** dari **{total_data}** desa")
        
        with tab5:
            st.markdown("### 💾 DOWNLOAD HASIL ETL DAN ANALISIS")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("#### 📊 Hasil ETL (Star Schema)")
                
                csv_fact = df_fact.to_csv(index=False)
                st.download_button(
                    label="📥 Download Fact Gizi Balita",
                    data=csv_fact,
                    file_name="fact_kesehatan.csv",
                    mime="text/csv",
                    use_container_width=True
                )
                
                csv_wilayah = df_wilayah.to_csv(index=False)
                st.download_button(
                    label="📥 Download Dimensi Wilayah",
                    data=csv_wilayah,
                    file_name="dim_wilayah.csv",
                    mime="text/csv",
                    use_container_width=True
                )
                
                csv_waktu = df_waktu.to_csv(index=False)
                st.download_button(
                    label="📥 Download Dimensi Waktu",
                    data=csv_waktu,
                    file_name="dim_waktu.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            with col2:
                st.markdown("#### 📈 Data Analisis")
                
                csv_agg = df_agg.to_csv(index=False)
                st.download_button(
                    label="📥 Download Data Agregat",
                    data=csv_agg,
                    file_name="data_agregat_puskesmas.csv",
                    mime="text/csv",
                    use_container_width=True
                )
                
                summary_data = {
                    'Indikator': ['Total Balita Ditimbang', 'Total Stunting', 'Persentase Stunting Rata-rata',
                                 'Total Kurang Gizi', 'Total Wasting', 'Jumlah Puskesmas'],
                    'Nilai': [total_ditimbang, total_stunting, f"{avg_stunting:.2f}%",
                             total_kurang_gizi, total_wasting, len(df_agg)]
                }
                df_summary = pd.DataFrame(summary_data)
                csv_summary = df_summary.to_csv(index=False)
                
                st.download_button(
                    label="📥 Download Ringkasan Statistik",
                    data=csv_summary,
                    file_name="ringkasan_statistik.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            st.markdown("---")
            
            st.markdown("""
            <div class="info-box">
                <b>ℹ️ Informasi File</b><br><br>
                📄 <b>Fact Kesehatan:</b> Tabel fakta berisi semua data gizi per puskesmas/desa<br>
                🏥 <b>Dimensi Wilayah:</b> Daftar puskesmas dan desa<br>
                📅 <b>Dimensi Waktu:</b> Informasi waktu pengambilan data<br>
                📊 <b>Data Agregat:</b> Ringkasan data per puskesmas (sudah diagregasi)<br>
                📈 <b>Ringkasan Statistik:</b> Statistik umum untuk laporan<br>
                🖼️ <b>Grafik PNG:</b> Tersedia tombol download di setiap grafik
            </div>
            """, unsafe_allow_html=True)

        # Footer dengan styling baru
        st.markdown("---")
        waktu_info_bawah = f"{df_waktu['tanggal'].iloc[0]} {df_waktu['bulan'].iloc[0]} {df_waktu['tahun'].iloc[0]}"
        st.markdown(f"""
        <div class="footer-card">
            <h3>📅 Data Terakhir Diperbarui</h3>
            <p style='font-size: 1.4rem; font-weight: 600; margin: 1rem 0;'>DATA STUNTING KABUPATEN KUNINGAN BULAN {pilih_bulan} </p>
            <p style='font-size: 1.4rem; font-weight: 600; margin: 1rem 0;'>(Waktu Penarikan : {waktu_info_bawah} )</p>
            <p style='font-size: 1.1rem;'>🏥 Dinas Kesehatan Kabupaten Kuningan</p>
            <p style='font-size: 0.95rem; opacity: 0.9;'>Sistem Informasi Analisis Data Stunting</p>
        </div>
        """, unsafe_allow_html=True)
    
    else:
        st.error(f"❌ {message}")
        st.markdown("""
        <div class="info-box">
            <b>🔧 Tips Troubleshooting</b><br><br>
            ✓ Pastikan upload 2 file: Status Gizi dan Sasaran Balita<br>
            ✓ Periksa format tanggal di baris pertama file<br>
            ✓ Pastikan data dimulai dari baris ke-4<br>
            ✓ Cek apakah semua kolom tersedia
        </div>
        """, unsafe_allow_html=True)